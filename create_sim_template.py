import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
NOTE_FILL = PatternFill("solid", fgColor="D9E1F2")
NOTE_FONT = Font(italic=True, color="1F4E78", name="Arial", size=9)

SHEETS = {
    "Atama": {
        "cols": ["Sicil", "Portfoy", "Portfoy Seviyesi"],
        "note": "O güne ait aktif atamalar. Portfoy Seviyesi: ANA / DESTEK / GECİCİ",
    },
    "Sicil_Hiz_Gun": {
        "cols": ["Portfoy", "Sicil", "Calisma_Suresi_Sn", "Referans_Adedi"],
        "note": "O gün sicil başına portföy bazında gerçek çalışma süresi (saniye) ve işlenen referans adedi.",
    },
    "Portfoy_Is_Yuku_Gun": {
        "cols": ["Portfoy", "Gelen_Ref"],
        "note": "O gün portföy bazında toplam gelen referans adedi.",
    },
    "Havuzda_Bekleme": {
        "cols": [
            "Portfoy", "Tarih", "Saat",
            "Gelen_Ref", "Ayni_Saatte_Baslanan",
            "Ort_Ilk_Temas_Sn", "Toplam_Calisilan_Ref", "Aktif_Sicil_Adedi",
        ],
        "note": "OPSİYONEL. Saatlik portföy bekleme verisi. Tarih: GG.AA.YYYY, Saat: HH:MM",
    },
}


def create_sim_template(path: str = "sample_data/simulasyon_sablon.xlsx"):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    dfs = {name: pd.DataFrame(columns=info["cols"]) for name, info in SHEETS.items()}
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in dfs.items():
            df.to_excel(writer, sheet_name=name, index=False, startrow=1)

    wb = load_workbook(path)
    for name, info in SHEETS.items():
        ws = wb[name]
        cols = info["cols"]
        ws.cell(row=1, column=1, value=info["note"])
        ws.cell(row=1, column=1).fill = NOTE_FILL
        ws.cell(row=1, column=1).font = NOTE_FONT
        ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        ws.row_dimensions[1].height = 28
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=2, column=ci, value=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 36
        for ci, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(col) * 0.9, 14)
    wb.save(path)


if __name__ == "__main__":
    create_sim_template()
    print("Simülasyon şablonu oluşturuldu: sample_data/simulasyon_sablon.xlsx")
