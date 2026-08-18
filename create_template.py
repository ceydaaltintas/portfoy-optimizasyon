import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
NOTE_FILL = PatternFill("solid", fgColor="D9E1F2")
NOTE_FONT = Font(italic=True, color="1F4E78", name="Arial", size=9)

SHEETS = {
    "Mevcut_Atama": {
        "cols": ["Sicil", "Portfoy", "Portfoy Seviyesi", "Baslangic Zamani", "Bitis Zamani"],
        "note": "Portfoy Seviyesi: ANA / DESTEK / GECİCİ  |  Baslangic/Bitis Zamani: yalnızca GECİCİ satırları için doldurulur (HH:MM formatı)",
    },
    "Portfoy_Is_Yuku": {
        "cols": [
            "Portfoy",
            "Toplam Referans Adedi",
            "Toplam Islem Adedi",
            "OM`ye Yonlendirilen Referans Adedi",
            "OM`ye Yonlendirilmeyen Referans Adedi",
            "Aktif Is Günü Sayisi",
            "Günlük Ortalama Referans Adedi",
            "Gunluk Medyan Referans Adedi",
            "Gunluk Ortalama Islem Adedi",
            "Gunluk Medyan Islem Adedi",
            "Gunluk Ortalama OM`ye Yonlendirilen Referans Adedi",
            "Gunluk Medyan OM`ye Yonlendirilen Referans Adedi",
            "Gunluk Ortalama OM`ye Yonlendirilmeyen Referans Adedi",
            "Gunluk Medyan OM`ye Yonlendirilmeyen Referans Adedi",
            "Gunluk Ortalama OM`ye Yonlendirilmeyen Islem Adedi",
            "Gunluk Medyan OM`ye Yonlendirilmeyen Islem Adedi",
        ],
        "note": "Her satır bir portföyü temsil eder. Toplam Islem Adedi >= Toplam Referans Adedi olmalıdır (1 referans = 1+ işlem).",
    },
    "Sicil_Hiz": {
        "cols": [
            "Portfoy",
            "Sicil",
            "Calistigi Is Gunu Sayisi",
            "Gunluk Ortalama Calisma Suresi",
            "Gunluk Medyan Calisma Suresi",
            "Gunluk Ortalama OM`ye yönlendirilen referanslarda calışma suresi",
            "Gunluk Medyan OM`ye yönlendirilen referanslarda calışma suresi",
            "Gunluk Ortalama OM`ye yönlendirilmeyen referanslarda calışma suresi",
            "Gunluk Medyan OM`ye yönlendirilmeyen referanslarda calışma suresi",
            "Gunluk Ortalama Referans Adedi",
            "Gunluk Medyan Referans Adedi",
        ],
        "note": "Her satır bir (Sicil, Portfoy) çiftini temsil eder. Tüm süreler saniye cinsindendir (sn/gün).",
    },
    "Portfoy_Aktif_Sicil": {
        "cols": [
            "Portfoy",
            "Gunluk Ortalama Aktif Sicil",
            "Gunluk Medyan Aktif Sicil",
            "Sicil bazlı günlük ortalama çalışma süresi",
            "Sicil bazlı günlük medyan çalışma süresi",
            "Sicil bazlı günlük ortalama OM`ye yonlendırılen referanslarda çalışma süresi",
            "Sicil bazlı günlük medyan OM`ye yonlendırılen referanslarda çalışma süresi",
            "Sicil bazlı günlük ortalama OM`ye yonlendırılmeyen referanslarda çalışma süresi",
            "Sicil bazlı günlük medyan OM`ye yonlendırılmeyen referanslarda çalışma süresi",
        ],
        "note": "Her satır bir portföyü temsil eder. Çalışma süreleri günlük toplam (sn/gün/sicil) cinsindedir.",
    },
    "Istisna": {
        "cols": ["Sicil", "Portfoy"],
        "note": "İki kullanım: (1) Sicil + Portfoy dolu → o çifte hiç atama yapılmaz. (2) Sadece Sicil dolu, Portfoy boş → o sicil optimizasyondan tamamen dışlanır.",
    },
    "Havuzda_Bekleme": {
        "cols": [
            "Portfoy",
            "Tarih",
            "Saat",
            "Gelen_Ref",
            "Ayni_Saatte_Baslanan",
            "Ort_Ilk_Temas_Sn",
            "Toplam_Calisilan_Ref",
            "Aktif_Sicil_Adedi",
        ],
        "note": "OPSİYONEL. Son 5 iş günü saatlik portföy verisi. Doldurulursa yoğun portföylerin talebi otomatik şişirilir. Tarih: YYYY-MM-DD, Saat: HH:MM, süreler saniye cinsinden.",
    },
}


def create_template(path: str = "sample_data/sablon.xlsx"):
    os.makedirs(os.path.dirname(path), exist_ok=True)

    dfs = {name: pd.DataFrame(columns=info["cols"]) for name, info in SHEETS.items()}
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in dfs.items():
            df.to_excel(writer, sheet_name=name, index=False, startrow=1)

    wb = load_workbook(path)
    for name, info in SHEETS.items():
        ws = wb[name]
        cols = info["cols"]

        # Not satırı (1. satır)
        ws.cell(row=1, column=1, value=info["note"])
        ws.cell(row=1, column=1).fill = NOTE_FILL
        ws.cell(row=1, column=1).font = NOTE_FONT
        ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        ws.row_dimensions[1].height = 30

        # Başlık satırı (2. satır)
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=2, column=ci, value=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 40

        # Kolon genişlikleri
        for ci, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(col) * 0.9, 14)

    wb.save(path)


if __name__ == "__main__":
    create_template()
    print("Şablon oluşturuldu: sample_data/sablon.xlsx")
