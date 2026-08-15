from __future__ import annotations
from datetime import date
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center")


def _style_sheet(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def export(sonuc: dict, data: dict) -> bytes:
    tum_siciller = data["tum_siciller"]
    ic_pf = data["ic_portfoyler"]
    gecici_pf = data["gecici_portfoyler"]
    tum_pf = ic_pf + gecici_pf
    capacity = data["capacity"]
    speed_raw = data["speed_raw"]
    istisna_set = data["istisna_set"]
    gecici_mevcut = data["gecici_mevcut"]

    ana_atama = sonuc["ana_atama"]       # sicil → portfoy
    destek_atama = sonuc["destek_atama"] # set[(sicil, portfoy)]
    demand = sonuc["demand"]
    ana_kapasite = sonuc["ana_kapasite"]
    destek_kapasite = sonuc["destek_kapasite"]
    coverage = sonuc["coverage"]
    uyarilar = sonuc["uyarilar"]

    # Ters map: portföy → [ANA siciller]
    pf_ana: dict[str, list] = {pf: [] for pf in ic_pf}
    for s, pf in ana_atama.items():
        if pf in pf_ana:
            pf_ana[pf].append(s)

    pf_destek: dict[str, list] = {pf: [] for pf in ic_pf}
    for s, pf in destek_atama:
        pf_destek[pf].append(s)

    wb = openpyxl.Workbook()

    # ── 1. Portföy Özet ───────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Portföy Özet"
    headers1 = ["Portföy", "Günlük Talep (sn)", "ANA Sicil Sayısı", "DESTEK Sicil Sayısı",
                "ANA Kapasite (sn)", "DESTEK Katkısı (sn)", "Toplam Kapasite (sn)",
                "Karşılama Oranı", "Ort. Hız Skoru (sn)"]
    ws1.append(headers1)

    for pf in ic_pf:
        ana_list = pf_ana.get(pf, [])
        dest_list = pf_destek.get(pf, [])
        all_sicil = ana_list + dest_list
        ort_hiz = sum(speed_raw.get(s, 0) for s in all_sicil) / len(all_sicil) if all_sicil else 0
        cov = coverage.get(pf, 0.0)
        ws1.append([
            pf,
            round(demand.get(pf, 0), 0),
            len(ana_list),
            len(dest_list),
            round(ana_kapasite.get(pf, 0), 0),
            round(destek_kapasite.get(pf, 0), 0),
            round(ana_kapasite.get(pf, 0) + destek_kapasite.get(pf, 0), 0),
            f"%{round(cov * 100, 1)}",
            round(ort_hiz, 0),
        ])

    # Karşılama oranı renklendirme
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        cov_cell = row[7]
        try:
            val = float(str(cov_cell.value).replace("%", ""))
            if val < 50:
                cov_cell.fill = PatternFill("solid", fgColor="FF0000")
                cov_cell.font = Font(name="Arial", size=10, color="FFFFFF", bold=True)
            elif val < 80:
                cov_cell.fill = PatternFill("solid", fgColor="FFA500")
                cov_cell.font = Font(name="Arial", size=10, bold=True)
        except Exception:
            pass
    _style_sheet(ws1)

    # ── 2. Sicil Özet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Sicil Özet")
    headers2 = ["Sicil", "ANA Portföy", "DESTEK Portföy Sayısı", "DESTEK Portföyler",
                "GECİCİ Portföy", "Kapasite (sn)", "Hız Skoru (sn)"]
    ws2.append(headers2)

    gecici_per_sicil: dict[str, list] = {}
    for g in gecici_mevcut:
        pf = g["portfoy"]
        lst = gecici_per_sicil.setdefault(g["sicil"], [])
        if pf not in lst:
            lst.append(pf)

    for s in tum_siciller:
        ana_pf = ana_atama.get(s, "—")
        dest_pf_list = [pf for (ss, pf) in destek_atama if ss == s]
        gecici_list = gecici_per_sicil.get(s, [])
        ws2.append([
            s,
            ana_pf,
            len(dest_pf_list),
            ", ".join(sorted(dest_pf_list)) if dest_pf_list else "—",
            ", ".join(gecici_list) if gecici_list else "—",
            round(capacity.get(s, 0), 0),
            round(speed_raw.get(s, 0), 0),
        ])
    _style_sheet(ws2)

    # ── 3. Atama Matrisi ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Atama Matrisi")
    ws3.append(["Sicil"] + tum_pf)

    gecici_set = set((g["sicil"], g["portfoy"]) for g in gecici_mevcut)

    for s in tum_siciller:
        row = [s]
        for pf in tum_pf:
            if ana_atama.get(s) == pf:
                row.append("ANA")
            elif (s, pf) in destek_atama:
                row.append("DESTEK")
            elif (s, pf) in gecici_set:
                row.append("GECİCİ")
            elif (s, pf) in istisna_set:
                row.append("İSTİSNA")
            else:
                row.append("—")
        ws3.append(row)

    # Renklendirme
    renkler = {"ANA": "1F4E78", "DESTEK": "2E75B6", "GECİCİ": "ED7D31", "İSTİSNA": "A6A6A6"}
    for row in ws3.iter_rows(min_row=2):
        for cell in row[1:]:
            c = renkler.get(str(cell.value), None)
            if c:
                cell.fill = PatternFill("solid", fgColor=c)
                cell.font = Font(name="Arial", size=9, color="FFFFFF", bold=(cell.value in ("ANA", "DESTEK")))
            else:
                cell.font = Font(name="Arial", size=9, color="AAAAAA")
            cell.alignment = CENTER
    _style_sheet(ws3)

    # ── 4. Detay Liste ────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Detay Liste")
    ws4.append(["Sicil", "Portföy", "Rol", "Kapasite (sn)", "Hız Skoru (sn)"])
    for s, pf in sorted(ana_atama.items()):
        ws4.append([s, pf, "ANA", round(capacity.get(s, 0), 0), round(speed_raw.get(s, 0), 0)])
    for s, pf in sorted(destek_atama):
        ws4.append([s, pf, "DESTEK", round(capacity.get(s, 0), 0), round(speed_raw.get(s, 0), 0)])
    for g in gecici_mevcut:
        ws4.append([g["sicil"], g["portfoy"], "GECİCİ",
                    round(capacity.get(g["sicil"], 0), 0), round(speed_raw.get(g["sicil"], 0), 0)])
    _style_sheet(ws4)

    # ── 5. Veri Uyarıları ────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Veri Uyarıları")
    ws5.append(["#", "Uyarı"])
    for i, u in enumerate(uyarilar, 1):
        ws5.append([i, u])
    _style_sheet(ws5)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
